# -*- coding: utf-8 -*-
"""
baixar_ensaios_api.py
=====================
Substituto do EXE "Extrair Ensaios AEVIAS" usando API em vez de Selenium.

DIFERENÇAS em relação ao EXE:
  - LISTA ensaios via API Base44 (sem Chrome, instantâneo)
  - BAIXA PDFs via Chrome CDP (igual ao EXE)
  - Output IDÊNTICO ao EXE: mesma pasta, mesma estrutura, mesmo Excel

Uso:
  python baixar_ensaios_api.py
  python baixar_ensaios_api.py --ini 01/04/2026 --fim 30/04/2026
  python baixar_ensaios_api.py --sem-pdf   (só lista + Excel)
"""
import os, sys, re, json, base64, socket, time, argparse, threading
from pathlib import Path
from datetime import datetime, timedelta

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

import tkinter as tk
from tkinter import messagebox

if sys.stdout:
    try: sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except Exception: pass

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

# =============================================================================
# CONFIGURAÇÕES — idêntico ao EXE
# =============================================================================
BASE_URL  = "https://aevias-controle.base44.app"
_APP_ID   = "68a7599ee3fb9205cfb852ec"
_API_BASE = f"{BASE_URL}/api/apps/{_APP_ID}/entities"

TOKEN = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9"
    ".eyJzdWIiOiJtYXRoZXVzLnJlc2VuZGVAYWZpcm1hZXZpYXMuY29tLmJyIiwiZXhwIjoxNzg0MjAxNzQzLCJpYXQiOjE3NzY0MjU3NDN9"
    ".nkIJL99iRM3asFuS3hv1XQMqK0aCUWr7CupxElwM15o"
)

# Detecta se é .py ou .exe (PyInstaller)
if getattr(sys, 'frozen', False):
    BASE_DIR = Path(sys.executable).parent
else:
    BASE_DIR = Path(__file__).parent

_output_env  = os.environ.get('AEVIAS_OUTPUT_DIR',  '').strip()
_backend_env = os.environ.get('AEVIAS_BACKEND_DIR', '').strip()
MODO_REDE    = os.environ.get('AEVIAS_MODO_REDE', '0') == '1'

OUTPUT_DIR  = Path(_output_env)  if _output_env  else BASE_DIR / "0.1-Resultados"
BACKEND_DIR = Path(_backend_env) if _backend_env else BASE_DIR / "0.2-BACKEND" / "Configuracoes_Internas"

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
BACKEND_DIR.mkdir(parents=True, exist_ok=True)

# Mesmo mapa de pastas do EXE
OBRA_FOLDER_MAP = {
    "SST":              "SST",
    "Pavimento":        "Pavimento",
    "TOPOGRAFIA":       "TOPOGRAFIA",
    "OAE / Terraplenos":"OAE - Terraplenos",
    "Ampliações":       "Ampliações",
    "ESCRITÓRIO":       "ESCRITÓRIO",
    "Conserva":         "Conserva",
}

# =============================================================================
# ENTIDADES BASE44
# =============================================================================
ENTIDADES = {
    "DiarioObra":                  ("Diário de Obra",          "/diario-de-obra",             "Pavimento"),
    "EnsaioCAUQ":                  ("Ensaio de CAUQ",          "/ensaio-cauq",                "Pavimento"),
    "ChecklistUsina":              ("Checklist de Usina",      "/checklist",                  "Pavimento"),
    "ChecklistAplicacao":          ("Checklist de Aplicação",  "/checklist-aplicacao",        "Pavimento"),
    "ChecklistMRAF":               ("Checklist de MRAF",       "/checklist-mraf",             "Pavimento"),
    "ChecklistTerraplanagem":      ("Checklist Terraplanagem", "/checklist-terraplanagem",    "OAE / Terraplenos"),
    "ChecklistConcretagem":        ("Checklist Concretagem",   "/checklist-concretagem",      "OAE / Terraplenos"),
    "ChecklistReciclagem":         ("Checklist Reciclagem",    "/checklist-reciclagem",       "Pavimento"),
    "AcompanhamentoUsinagem":      ("Acomp. Usinagem",         "/acompanhamento-usinagem",    "Pavimento"),
    "AcompanhamentoCarga":         ("Acomp. Carga",            "/acompanhamento-carga",       "Pavimento"),
    "EnsaioDensidadeInSitu":       ("Ensaio Densidade In Situ","/ensaio-densidade",           "Pavimento"),
    "EnsaioGranulometriaIndividual":("Ensaio Granulometria",   "/ensaio-granulometria",       "Pavimento"),
    "EnsaioManchaPendulo":         ("Ensaio Mancha/Pêndulo",   "/ensaio-mancha-pendulo",      "Pavimento"),
    "EnsaioVigaBenkelman":         ("Ensaio Viga Benkelman",   "/ensaio-viga-benkelman",      "Pavimento"),
    "EnsaioProctor":               ("Ensaio Proctor",          "/EnsaioProctor",              "OAE / Terraplenos"),
    "EnsaioTaxaMRAF":              ("Ensaio Taxa MRAF",        "/ensaio-taxa-mraf",           "Pavimento"),
    "EnsaioTaxaPinturaImprimacao": ("Ensaio Taxa Pintura",     "/ensaio-taxa-pintura",        "Pavimento"),
    "EnsaioSondagem":              ("Ensaio Sondagem",         "/ensaio-sondagem",            "OAE / Terraplenos"),
}

# =============================================================================
# HELPERS
# =============================================================================

def _headers():
    return {"Authorization": f"Bearer {TOKEN}", "Content-Type": "application/json"}


def sanitize(name: str) -> str:
    name = name.replace("/", "-").replace("\\", "-")
    return re.sub(r'[<>:"|?*]', '', name).strip(". ")


def _get_lab(rec: dict) -> str:
    nome = rec.get("laboratorista_name", "")
    if nome: return nome
    cb = rec.get("created_by")
    if isinstance(cb, dict): return cb.get("full_name", "—")
    return str(cb or "—")


def _data_fmt(rec: dict) -> str:
    raw = rec.get("data") or rec.get("data_ensaio") or rec.get("created_date", "")
    try:
        return datetime.fromisoformat(str(raw)[:10]).strftime("%d/%m/%Y")
    except Exception:
        return str(raw)[:10]


def _status(rec: dict) -> str:
    if rec.get("was_rejected"): return "Reprovado"
    if rec.get("approved"):     return "Aprovado"
    return "Pendente"

# =============================================================================
# INTERFACE TKINTER (igual ao EXE)
# =============================================================================

def obter_intervalo_datas():
    ini_env = os.environ.get('AEVIAS_DATA_INI', '').strip()
    fim_env = os.environ.get('AEVIAS_DATA_FIM', '').strip()
    if ini_env and fim_env:
        try:
            return {"inicio": datetime.strptime(ini_env, "%d/%m/%Y"),
                    "fim":    datetime.strptime(fim_env, "%d/%m/%Y"),
                    "cancelado": False}
        except ValueError:
            pass

    resultado = {"inicio": None, "fim": None, "cancelado": True}
    root = tk.Tk()
    root.title("Seleção de Período - AEVIAS (API)")
    root.geometry("350x250")
    root.resizable(False, False)
    root.attributes("-topmost", True)
    root.lift(); root.focus_force()
    sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
    root.geometry(f"+{sw//2-175}+{sh//2-125}")

    tk.Label(root, text="Selecione o período dos ensaios", font=("Arial", 12, "bold")).pack(pady=10)
    tk.Label(root, text="Data Inicial (DD/MM/AAAA):", font=("Arial", 10)).pack()
    e_ini = tk.Entry(root, justify='center', font=("Arial", 10))
    e_ini.insert(0, datetime.now().strftime("01/%m/%Y")); e_ini.pack(pady=5)
    tk.Label(root, text="Data Final (DD/MM/AAAA):", font=("Arial", 10)).pack()
    e_fim = tk.Entry(root, justify='center', font=("Arial", 10))
    e_fim.insert(0, datetime.now().strftime("%d/%m/%Y")); e_fim.pack(pady=5)

    def confirmar():
        try:
            di = datetime.strptime(e_ini.get().strip(), "%d/%m/%Y")
            df = datetime.strptime(e_fim.get().strip(), "%d/%m/%Y")
            if di > df:
                messagebox.showerror("Erro", "Data inicial maior que final!"); return
            resultado.update({"inicio": di, "fim": df, "cancelado": False})
            root.destroy()
        except ValueError:
            messagebox.showerror("Erro", "Formato inválido! Use DD/MM/AAAA")

    frm = tk.Frame(root); frm.pack(pady=15)
    tk.Button(frm, text="Confirmar", command=confirmar,
              bg="#4CAF50", fg="white", font=("Arial", 10, "bold"), width=12).pack(side=tk.LEFT, padx=5)
    tk.Button(frm, text="Cancelar", command=root.destroy,
              bg="#f44336", fg="white", font=("Arial", 10, "bold"), width=12).pack(side=tk.LEFT, padx=5)
    root.mainloop()
    return resultado

# =============================================================================
# MAPA DE OBRAS (API)
# =============================================================================

def _carregar_mapa_obras() -> dict:
    print("  Carregando mapa de obras via API...")
    try:
        r = requests.get(f"{_API_BASE}/Obra", headers=_headers(), timeout=30)
        obras = r.json() if r.status_code == 200 else []
    except Exception:
        return {}

    def _grupo(nome):
        n = (nome or "").upper()
        if "SST" in n or "SEGURANÇA" in n: return "SST"
        if "TOPO" in n:                    return "TOPOGRAFIA"
        if "ESCRIT" in n:                  return "ESCRITÓRIO"
        if "OAE" in n or "TERRAPLAN" in n: return "OAE / Terraplenos"
        if "CONSERVA" in n:                return "Conserva"
        if "AMPLIA" in n:                  return "Ampliações"
        return "Pavimento"

    return {o["id"]: _grupo(o.get("name", "")) for o in obras if o.get("id")}

# =============================================================================
# LISTAGEM VIA API (substitui o scraping Selenium)
# =============================================================================

def listar_todos(data_ini: datetime, data_fim: datetime) -> list[dict]:
    mapa_obras = _carregar_mapa_obras()
    resultado  = []

    print(f"\n  Listando via API ({data_ini.strftime('%d/%m/%Y')} → {data_fim.strftime('%d/%m/%Y')})...")

    for entidade, (tipo_nome, path, obra_default) in ENTIDADES.items():
        try:
            r = requests.get(f"{_API_BASE}/{entidade}", headers=_headers(), timeout=30)
            if r.status_code != 200:
                print(f"    [WARN] {entidade}: HTTP {r.status_code}")
                continue
            recs = r.json()
            if not isinstance(recs, list): continue
        except Exception as e:
            print(f"    [ERR] {entidade}: {e}"); continue

        count = 0
        for rec in recs:
            data_str = _data_fmt(rec)
            try:
                dt = datetime.strptime(data_str, "%d/%m/%Y")
                if not (data_ini <= dt <= data_fim): continue
            except Exception:
                continue

            lab    = _get_lab(rec)
            rec_id = rec.get("id", "")
            oid    = rec.get("obra_id") or rec.get("project_id") or ""
            obra   = mapa_obras.get(oid, obra_default)

            resultado.append({
                "tipo":        tipo_nome,
                "profissional": lab,
                "lab":         lab,
                "data":        data_str,
                "obra":        obra,
                "contrato":    "",
                "local":       rec.get("trecho", ""),
                "empreiteira": "",
                "projeto":     "",
                "status":      _status(rec),
                "reportUrl":   f"{path}/{rec_id}" if rec_id else path,
                "id":          rec_id,
            })
            count += 1

        if count:
            print(f"    ✓ {entidade}: {count} registros")

    print(f"\n  Total API: {len(resultado)} registros")
    return resultado

# =============================================================================
# CHROME / CDP  (igual ao EXE)
# =============================================================================

def setup_driver():
    options = Options()
    machine_id = re.sub(r'[^a-zA-Z0-9_-]', '_', socket.gethostname())
    profile_dir = BACKEND_DIR / ".chrome_cache" / machine_id
    profile_dir.mkdir(parents=True, exist_ok=True)
    options.add_argument(f"--user-data-dir={profile_dir}")
    options.add_argument("--profile-directory=Default")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--start-maximized")
    prefs = {
        "printing.print_preview_sticky_settings.appState": json.dumps({
            "recentDestinations": [{"id": "Save as PDF", "origin": "local", "account": ""}],
            "selectedDestinationId": "Save as PDF", "version": 2
        }),
        "savefile.default_directory": str(OUTPUT_DIR),
        "download.default_directory": str(OUTPUT_DIR),
        "download.prompt_for_download": False,
    }
    options.add_experimental_option("prefs", prefs)
    options.add_argument("--kiosk-printing")
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)


def aguardar_login(driver) -> str:
    """Abre o site, aguarda login e retorna nome do usuário (igual ao EXE)."""
    driver.get(f"{BASE_URL}/MeusEnsaios")
    time.sleep(6)

    def capturar_usuario():
        time.sleep(4)
        seletores = [
            "//header//button//div[string-length(text()) > 0]",
            "//button[contains(@class, 'user')]//span",
            "//span[contains(@class, 'user-name')]",
            "//*[contains(text(), '@')]",
        ]
        for sel in seletores:
            try:
                el = driver.find_element(By.XPATH, sel)
                txt = el.text.strip()
                if txt and len(txt) > 1 and "Sair" not in txt:
                    if "@" in txt: txt = txt.split("@")[0]
                    return sanitize(txt.replace("Olá,", "").replace("Olá", "").strip())
            except Exception:
                continue
        return sanitize(os.environ.get('USERNAME', 'Usuario_AEvias'))

    try:
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Ensaios Realizados')]"))
        )
        print("Já logado!")
        return capturar_usuario()
    except Exception:
        pass

    print("\n" + "=" * 50)
    print("  FAÇA LOGIN NO CHROME QUE ABRIU")
    print("  O script detectará automaticamente...")
    print("=" * 50)

    for attempt in range(60):
        time.sleep(5)
        try:
            driver.find_element(By.XPATH, "//*[contains(text(), 'Ensaios Realizados')]")
            print("  Login confirmado!")
            return capturar_usuario()
        except Exception:
            pass
        current = driver.current_url
        if "login" not in current and "MeusEnsaios" not in current:
            driver.get(f"{BASE_URL}/MeusEnsaios")
            time.sleep(5)

    return "Usuario_AEvias"


def gerar_pdf_cdp(driver, url: str, output_path: Path, timeout: int = 45) -> bool:
    result: dict = {"data": None, "error": None}

    def do_pdf():
        try:
            driver.set_page_load_timeout(timeout)
            driver.get(url)
            time.sleep(3)
            try:
                WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH,
                    "//*[contains(text(),'GERAIS') or contains(text(),'Relat') or "
                    "contains(text(),'Checklist') or contains(text(),'Ensaio') or "
                    "contains(text(),'Resultado')]"
                )))
            except Exception:
                pass
            # Aguarda "Otimizando imagens"
            ot = "//*[contains(text(),'Otimizando imagens')]"
            try:
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, ot)))
                print("    Otimizando imagens... aguardando.")
                WebDriverWait(driver, 60).until_not(EC.presence_of_element_located((By.XPATH, ot)))
            except Exception:
                pass
            time.sleep(2)
            r = driver.execute_cdp_cmd("Page.printToPDF", {
                "printBackground": True, "preferCSSPageSize": True,
                "marginTop": 0.4, "marginBottom": 0.4,
                "marginLeft": 0.4, "marginRight": 0.4,
                "paperWidth": 8.27, "paperHeight": 11.69,
            })
            result["data"] = r["data"]
        except Exception as e:
            result["error"] = str(e)

    t = threading.Thread(target=do_pdf)
    t.start(); t.join(timeout + 5)

    if t.is_alive():
        try: driver.execute_script("window.stop();")
        except Exception: pass
        return False

    if result.get("error") or not result.get("data"):
        return False

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "wb") as f:
        f.write(base64.b64decode(result["data"]))
    return True

# =============================================================================
# DOWNLOAD DE PDFs (estrutura idêntica ao EXE)
# =============================================================================

def download_all_pdfs(driver, ensaios: list[dict], org_root: Path):
    total = len(ensaios); ok = 0; erros = []

    for i, ens in enumerate(ensaios, 1):
        obra       = ens.get("obra", "Sem Obra")
        lab        = ens.get("lab", "Sem Colaborador")
        tipo       = ens.get("tipo", "Sem Tipo")
        data       = ens.get("data", "Sem Data").replace("/", "-")
        profissional = ens.get("profissional", lab)
        report_url = ens.get("reportUrl", "")

        if not report_url or report_url == "#":
            erros.append(f"Sem URL: {tipo} | {lab}")
            continue

        # Pasta: Obra / Colaborador  (igual ao EXE)
        obra_folder  = sanitize(OBRA_FOLDER_MAP.get(obra, obra))
        colab_folder = sanitize(lab)
        target_dir   = org_root / obra_folder / colab_folder
        target_dir.mkdir(parents=True, exist_ok=True)

        # Nome: Tipo - DD-MM-AAAA - Profissional.pdf  (igual ao EXE)
        nome_display = profissional if (profissional and profissional != "Sem Profissional") else lab
        filename = sanitize(f"{tipo} - {data} - {nome_display}.pdf")
        output_path = target_dir / filename

        # Evita duplicata
        counter = 2
        while output_path.exists():
            stem = filename.rsplit(".pdf", 1)[0]
            output_path = target_dir / f"{stem} ({counter}).pdf"
            counter += 1

        full_url = f"{BASE_URL}{report_url}"
        print(f"  [{i}/{total}] {obra_folder} > {colab_folder} > {filename[:50]}")

        try:
            if gerar_pdf_cdp(driver, full_url, output_path):
                ok += 1
            else:
                erros.append(f"Falha CDP: {filename}")
        except Exception as e:
            erros.append(f"{filename}: {e}")

    return ok, erros

# =============================================================================
# RELATÓRIO EXCEL (idêntico ao EXE)
# =============================================================================

def gerar_relatorio_presenca(ensaios: list[dict], data_ini: datetime,
                              data_fim: datetime, output_path: Path):
    print(f"\n  Gerando relatório: {output_path.name}...")
    df = pd.DataFrame(ensaios)
    if df.empty: return

    df['dt'] = df['data'].apply(lambda x: datetime.strptime(x.split()[0], "%d/%m/%Y"))

    def categorizar(tipo):
        t = tipo.upper()
        if "CHECKLIST" in t: return "Checklist"
        if "DIÁRIO" in t or "DIARIO" in t: return "Diário de Obra"
        return "Ensaio"
    df['categoria'] = df['tipo'].apply(categorizar)

    dias = []
    curr = data_ini
    while curr <= data_fim:
        dias.append(curr); curr += timedelta(days=1)

    wb = Workbook()
    del wb['Sheet']

    fill_green = PatternFill("solid", fgColor="C6EFCE")
    fill_gray  = PatternFill("solid", fgColor="F2F2F2")
    fill_hdr_a = PatternFill("solid", fgColor="1A3A5C")
    fill_sub   = PatternFill("solid", fgColor="2E6DA4")
    fill_col   = PatternFill("solid", fgColor="4472C4")
    fill_colab = PatternFill("solid", fgColor="EBF3FB")
    f_ok       = Font(color="006100", bold=True)
    f_muted    = Font(color="808080")
    f_hdr      = Font(bold=True, color="FFFFFF")
    f_hdr10    = Font(size=10, color="FFFFFF")
    f_hdr_b    = Font(bold=True, color="FFFFFF")
    alc        = Alignment(horizontal="center", vertical="center")
    brd        = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

    n_dias = len(dias)
    n_cols = 2 + n_dias
    cats   = ["Checklist", "Diário de Obra", "Ensaio"]
    collabs = sorted(df['lab'].unique())

    # ── Aba Resumão (primeira) ───────────────────────────────────────────────
    ws_r = wb.create_sheet("Resumão", 0)
    ws_r.merge_cells(1, 1, 1, n_cols)
    c = ws_r.cell(1, 1, "RESUMÃO — ENSAIOS AEVIAS")
    c.font = Font(bold=True, size=14, color="FFFFFF"); c.fill = fill_hdr_a
    c.alignment = alc; ws_r.row_dimensions[1].height = 30

    ws_r.merge_cells(2, 1, 2, n_cols)
    c = ws_r.cell(2, 1, f"Período: {data_ini.strftime('%d/%m/%Y')} → {data_fim.strftime('%d/%m/%Y')}  |  Total: {len(df)}")
    c.font = f_hdr10; c.fill = fill_sub; c.alignment = alc

    r = 4
    for ci, h in enumerate(["COLABORADOR", "CATEGORIA"], 1):
        c = ws_r.cell(r, ci, h); c.font = f_hdr_b; c.fill = fill_col
        c.alignment = alc; c.border = brd
    for i, d in enumerate(dias, 3):
        c = ws_r.cell(r, i, d.strftime("%d/%m")); c.font = f_hdr_b; c.fill = fill_col
        c.alignment = alc; c.border = brd
        ws_r.column_dimensions[get_column_letter(i)].width = 6
    ws_r.row_dimensions[r].height = 18; r += 1

    for colab in collabs:
        df_c = df[df['lab'] == colab]; first = r
        for ci, cat in enumerate(cats):
            c = ws_r.cell(r, 2, cat)
            c.font = Font(bold=(ci == 0))
            c.alignment = Alignment(horizontal="left", vertical="center"); c.border = brd
            if ci == 0: c.fill = fill_colab
            for col_i, d in enumerate(dias, 3):
                dom  = (d.weekday() == 6)
                teve = not df_c[(df_c['dt'] == d) & (df_c['categoria'] == cat)].empty
                cel  = ws_r.cell(r, col_i); cel.alignment = alc; cel.border = brd
                if teve:
                    cel.value = "OK"; cel.fill = fill_green; cel.font = f_ok
                else:
                    cel.value = "-"; cel.font = f_muted
                    if dom: cel.fill = fill_gray
                    elif ci == 0: cel.fill = fill_colab
            ws_r.row_dimensions[r].height = 15; r += 1
        ws_r.merge_cells(first, 1, r - 1, 1)
        c = ws_r.cell(first, 1, colab)
        c.font = Font(bold=True, size=10); c.fill = fill_colab
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = brd
    ws_r.column_dimensions['A'].width = 24
    ws_r.column_dimensions['B'].width = 18

    # ── Abas individuais por colaborador ─────────────────────────────────────
    for colab in collabs:
        sn   = re.sub(r'[\\*?:/\[\]]', '', colab)[:31]
        ws2  = wb.create_sheet(sn)
        df_c = df[df['lab'] == colab]
        c = ws2.cell(1, 1, "CATEGORIA"); c.font = f_hdr; c.fill = fill_col; c.alignment = alc
        for i, d in enumerate(dias, 2):
            c = ws2.cell(1, i, d.strftime("%d/%m")); c.font = f_hdr; c.fill = fill_col
            c.alignment = alc
            ws2.column_dimensions[get_column_letter(i)].width = 6
        ws2.column_dimensions['A'].width = 20
        for ri, cat in enumerate(cats, 2):
            ws2.cell(ri, 1, cat).font = Font(bold=True); ws2.cell(ri, 1).border = brd
            for ci, d in enumerate(dias, 2):
                dom  = (d.weekday() == 6)
                teve = not df_c[(df_c['dt'] == d) & (df_c['categoria'] == cat)].empty
                cel  = ws2.cell(ri, ci); cel.border = brd; cel.alignment = alc
                if teve:
                    cel.value = "OK"; cel.fill = fill_green; cel.font = f_ok
                else:
                    cel.value = "-"; cel.font = f_muted
                    if dom: cel.fill = fill_gray

    wb.save(output_path)
    print(f"  ✓ Relatório salvo: {output_path}")

# =============================================================================
# MAIN
# =============================================================================

def parse_args():
    p = argparse.ArgumentParser(description="Baixar ensaios AEVIAS via API")
    p.add_argument("--ini",      default="", help="Data inicial DD/MM/AAAA")
    p.add_argument("--fim",      default="", help="Data final DD/MM/AAAA")
    p.add_argument("--sem-pdf",  action="store_true", help="Só lista + Excel")
    return p.parse_args()


def main():
    args = parse_args()

    # Período via Tkinter (ou args)
    if args.ini and args.fim:
        try:
            data_ini = datetime.strptime(args.ini, "%d/%m/%Y")
            data_fim = datetime.strptime(args.fim, "%d/%m/%Y")
        except ValueError:
            print("Formato inválido. Use DD/MM/AAAA"); return
    else:
        periodo = obter_intervalo_datas()
        if periodo["cancelado"]:
            print("\nOperação cancelada."); return
        data_ini = periodo["inicio"]
        data_fim = periodo["fim"]

    timestamp = datetime.now().strftime("%d-%m-%Y_%H-%M")

    print("=" * 60)
    print("  AEVIAS CONTROLE — Download via API")
    print(f"  Período: {data_ini.strftime('%d/%m/%Y')} → {data_fim.strftime('%d/%m/%Y')}")
    print("=" * 60)

    # ── ETAPA 1: Listar via API ───────────────────────────────────────────────
    print("\n" + "=" * 40)
    print("  ETAPA 1: Listando ensaios via API Base44")
    print("=" * 40)
    ensaios = listar_todos(data_ini, data_fim)

    if not ensaios:
        print("Nenhum ensaio encontrado no período."); return

    # Nome de usuário (sem Chrome por enquanto)
    nome_usuario = sanitize(os.environ.get('USERNAME', 'matheus.resende'))

    # Pasta raiz (mesmo padrão do EXE)
    pasta_raiz_nome = f"Ensaios_{nome_usuario}_{timestamp}"
    if MODO_REDE:
        periodo_str = f"{data_ini.strftime('%d-%m-%Y')}_a_{data_fim.strftime('%d-%m-%Y')}"
        ORG_ROOT = OUTPUT_DIR / nome_usuario / periodo_str
    else:
        ORG_ROOT = OUTPUT_DIR / pasta_raiz_nome
    ORG_ROOT.mkdir(parents=True, exist_ok=True)

    print(f"\n  Pasta destino: {ORG_ROOT}")

    # Grava ultima_pasta.txt (compatibilidade com executar_tudo.py)
    ultima_pasta_file = BACKEND_DIR / "ultima_pasta.txt"
    ultima_pasta_file.write_text(str(ORG_ROOT), encoding='utf-8')

    # ── JSON backup (mesmo nome do EXE) ──────────────────────────────────────
    json_path = ORG_ROOT / "ensaios_dados_filtrados.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(ensaios, f, ensure_ascii=False, indent=2)
    print(f"  JSON salvo: {json_path}")

    # ── ETAPA 2: Baixar PDFs ─────────────────────────────────────────────────
    ok = 0; erros = []

    if not args.sem_pdf:
        print("\n" + "=" * 40)
        print("  ETAPA 2: Baixando PDFs via Chrome CDP")
        print("=" * 40)
        print(f"  Total a baixar: {len(ensaios)}\n")

        driver = None
        try:
            driver = setup_driver()
            # Captura nome do usuário logado (igual ao EXE)
            nome_logado = aguardar_login(driver)
            if nome_logado and nome_logado != "Usuario_AEvias":
                # Renomeia pasta para usar o nome real do usuário logado
                novo_nome = f"Ensaios_{nome_logado}_{timestamp}"
                nova_root = OUTPUT_DIR / novo_nome
                ORG_ROOT.rename(nova_root)
                ORG_ROOT = nova_root
                ultima_pasta_file.write_text(str(ORG_ROOT), encoding='utf-8')
                print(f"  Pasta renomeada: {ORG_ROOT}")

            ok, erros = download_all_pdfs(driver, ensaios, ORG_ROOT)
        finally:
            if driver:
                print("\nFechando Chrome em 3 segundos...")
                time.sleep(3); driver.quit()
    else:
        print("\n  Modo --sem-pdf: pulando download de PDFs.")

    # ── ETAPA 3: Excel de presença ────────────────────────────────────────────
    print("\n" + "=" * 40)
    print("  ETAPA 3: Gerando Relatório Visual")
    print("=" * 40)
    excel_path = ORG_ROOT / f"Relatorio_Atividades_{nome_usuario}_{timestamp}.xlsx"
    try:
        gerar_relatorio_presenca(ensaios, data_ini, data_fim, excel_path)
    except Exception as e:
        print(f"  ERRO Excel: {e}")

    # ── Relatório final ───────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("  RELATÓRIO FINAL")
    print("=" * 60)
    print(f"  Total de ensaios : {len(ensaios)}")
    print(f"  PDFs baixados    : {ok}")
    print(f"  Erros            : {len(erros)}")
    if erros:
        print("\n  ERROS:")
        for err in erros: print(f"    - {err}")

    print("\n  ESTRUTURA DE PASTAS:")
    if ORG_ROOT.exists():
        for obra_dir in sorted(ORG_ROOT.iterdir()):
            if obra_dir.is_dir():
                pdfs = list(obra_dir.glob("**/*.pdf"))
                if pdfs:
                    print(f"\n    [{obra_dir.name}] ({len(pdfs)} PDFs)")
                    for colab_dir in sorted(obra_dir.iterdir()):
                        if colab_dir.is_dir():
                            n = len(list(colab_dir.glob("*.pdf")))
                            if n: print(f"      └── {colab_dir.name}: {n} PDFs")

    print(f"\n  ✓ Concluído! Arquivos em: {ORG_ROOT}")


if __name__ == "__main__":
    main()
